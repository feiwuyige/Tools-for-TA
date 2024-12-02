import openpyxl
from openpyxl.styles import PatternFill
import os

# 读取作业文件列表
def get_homework_list(directory):
    homework_list = []
    for filename in os.listdir(directory):
            # 防止有些同学命名有错误   
            if len(filename) >= 12:
                homework_list.append(filename[-12 : ])
    return homework_list

# 判断作业提交情况
def check_homework(homework_list, min_row, max_row):
    # 加载excel文件
    file_path = r'your_path'
    workbook = openpyxl.load_workbook(file_path)
    
    # 选择工作表
    sheet = workbook['Sheet1']

    # 学号在B列
    column_letter = 'B'
    # 要赋值的单元格在 C 列
    assignment_letter = 'C'

    # 不符合命名的作业
    error_name = []

    # 遍历homework_list列表
    for homework in homework_list:
        flag = False
        for row in range(min_row, max_row + 1):
            cell_value = sheet[f'{column_letter}{row}'].value
            if cell_value == homework:
                sheet[f'{assignment_letter}{row}'] = 1
                flag = True
                break
        if not flag:
            error_name.append(homework)
    # 保存
    workbook.save(file_path)
    print("作业检查完毕, 不符合命名要求的有：")
    for name in error_name:
        print(name)


# 最后处理excel表格，将未提交的置为0，并标黄
def handle_not_submit(min_row, max_row):
     # 加载excel文件
    file_path = r'your_path'
    workbook = openpyxl.load_workbook(file_path)
    
    # 选择工作表
    sheet = workbook['Sheet1']

    # 要赋值的单元格在 C 列
    assignment_letter = 'C'

    # 创建一个黄色填充样式
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    # 遍历homework_list列表
    for row in range(min_row, max_row + 1):
        cell_value = sheet[f'{assignment_letter}{row}'].value
        if cell_value == None:
            sheet[f'{assignment_letter}{row}'] = 0
            for cell in sheet[row]:
                cell.fill = yellow_fill
            
    # 保存
    workbook.save(file_path)
    print("最后处理完毕")
                   

if __name__ == '__main__':
    homework_directory = r'your_path'
    homework_list = get_homework_list(homework_directory)
    print(f'共有作业{len(homework_list)}份')
    min_row = 1
    max_row = 88 # 根据excel表格调整
    check_homework(homework_list, min_row, max_row)
    handle_not_submit(min_row, max_row)
